import openpyxl

workbook=openpyxl.Workbook()
sheet = workbook.active
data = [["Name","Age","City"],
        ["sundaram",19,"trilokpur"],
        ["sundaram",19,"trilokpur"],
        ["sundaram",19,"trilokpur"],
        ["sundaram",19,"trilokpur"]
        ]
for row in data:
    sheet.append(row)

workbook.save("file_excel.xlsx")

loaded_workbook = openpyxl.load_workbook("file_excel.xlsx")
loaded_sheet=loaded_workbook.active

for row in loaded_sheet.iter_rows(values_only=True):
    print(row)
